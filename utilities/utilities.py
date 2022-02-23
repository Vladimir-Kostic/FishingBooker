import inspect
import logging
from openpyxl import load_workbook


class Utils:
    def logging_setup(log_level=logging.DEBUG):
        log_name = inspect.stack()[1][3]
        logger = logging.getLogger(log_name)
        logger.setLevel(log_level)
        file_handler = logging.FileHandler(".\\logs\\log_report.log")
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(name)s : %(message)s', datefmt='%d/%m/%Y %I:%M:%S %p')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        return logger

    def read_data_from_excel(file_name, sheet):
        data_list = []
        wb = load_workbook(file_name)
        sheet = wb[sheet]
        row_count = sheet.max_row
        column_count = sheet.max_column
        for i in range(2, row_count + 1):
            rows = []
            for j in range(1, column_count + 1):
                rows.append(sheet.cell(row=i, column=j).value)
            data_list.append(rows)
        return data_list

    def write_text_to_file(file_name, text):
        with open("Verification_ID.txt", "w") as file:
            file.write(text)
            return file
